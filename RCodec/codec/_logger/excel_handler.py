import os
import re
from copy import copy
from typing import Union, Optional, Sequence, Iterable, Dict

import openpyxl as xl
from openpyxl.utils.cell import column_index_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl.writer.excel import save_workbook

from ..common import Mode, PatKey
from .record import Record
from .resource import TEMPLATE


class Excel(object):
    def __init__(self, name: str):
        self.name = name
        if not os.path.exists(name):
            save_workbook(Workbook(), name)
        self._wb = xl.load_workbook(name, keep_vba=True)

    def __getitem__(self, item: Union[str, tuple, int]) -> Worksheet:
        """
        如果已存在，则直接返回，否则如果不存在，则创建
        :param item:
        :return:
        """
        if isinstance(item, str):
            if item in self._wb:
                return self._wb[item]
            else:
                return self._wb.create_sheet(item)
        elif isinstance(item, tuple):
            index, item = item
            if item in self._wb:
                return self._wb[item]
            else:
                return self._wb.create_sheet(item, index=index)
        elif isinstance(item, int):
            return self._wb.worksheets[item]
        else:
            raise ValueError("unsupported key")

    def __contains__(self, item):
        if isinstance(item, Mode):
            item = item.value
        return item in self._wb

    def __len__(self):
        return len(self._wb.sheetnames)

    def delete(self, sheet: Worksheet):
        """
        删除指定的sheet
        :param sheet: sheet对象
        """
        self._wb.remove(sheet)

    def save(self, name: Optional[str] = None):
        """
        保存当前Workbook
        :param name: 保存到目标文件名称
        """
        if name is None:
            name = self.name
        self._wb.save(name)

    def close(self):
        """
        关闭Excel
        """
        self._wb.close()

    def copy_sheet(self, source: Worksheet, new_title: str, index: Optional[int] = None):
        target = self.__getitem__((index, new_title))
        WorksheetCopy(source_worksheet=source, target_worksheet=target).copy_worksheet()
        return target

    @staticmethod
    def from_template(seqs: Sequence[str], qps: Sequence[int], mode: Mode):
        target = Excel(TEMPLATE)
        # 填充原始数据的表头
        ExcelHelper.fill_raw_data_header(Mode, seqs, qps, target)
        # 填充具体表格的表头
        template_sheet = target[_RAW_DATA_SHEETS[2]]
        ExcelHelper.fill_ref_data_header(mode, seqs, qps, target, template_sheet, index=0)
        target.delete(template_sheet)

        return target

    def new_mode_sheet(self, seqs: Sequence[str], qps: Sequence[int], mode: Mode):
        index = len(self) - 2
        ExcelHelper.fill_ref_data_header(mode, seqs, qps, self, self[0], index=index)


_RAW_DATA_SHEETS = xl.load_workbook(TEMPLATE).sheetnames
_RAW_DATA_COL = ['B', 'D']


class ExcelHelper(object):
    @staticmethod
    def is_excel_file(filename: str):
        # copy from openpyxl.reader.excel.SUPPORTED_FORMATS
        supported_formats = ('.xlsx', '.xlsm', '.xltx', '.xltm')
        for fmt in supported_formats:
            if filename.endswith(fmt):
                return True
        return False

    @staticmethod
    def seq_id(seqs: Sequence[str], file: str) -> Optional[int]:
        """
        判断给定的file名称是否在初始化时的序列中
        :param seqs: 待扫描的所有序列
        :param file: excel 表格中的序列名称
        :return: 如果存在，则返回索引，否则，返回None
        """
        # 1. match name exactly
        for i, value in enumerate(seqs):
            if value in file.split("_") or value in "_".join(file.split("_")[1:3]):
                return i
        # 2. containing is ok
        for i, value in enumerate(seqs):
            if value.lower() in file.lower() or file.lower() in value.lower():
                return i
        # failed
        return None

    @staticmethod
    def fill_raw_data_header(modes: Iterable[Mode], seqs: Sequence[str], qps: Sequence[int], excel: Excel):
        sheet0 = excel[_RAW_DATA_SHEETS[0]]
        sheet1 = excel[_RAW_DATA_SHEETS[1]]
        row = 1
        col = column_index_from_string(_RAW_DATA_COL[0])
        for mode in modes:
            for seq in seqs:
                for qp in qps:
                    sheet0.cell(row=row, column=col, value=f"{mode.value.lower()}_{seq}_{qp}")
                    sheet1.cell(row=row, column=col, value=f"{mode.value.lower()}_{seq}_{qp}")
                    row += 1

    @staticmethod
    def fill_raw_data(target: Excel, is_anchor: bool, records: Dict[int, Iterable[Record]],
                      seqs: Sequence[str], qps: Sequence[int]):
        sheet: Worksheet = target[_RAW_DATA_SHEETS[0] if is_anchor else _RAW_DATA_SHEETS[1]]
        col_start = column_index_from_string(_RAW_DATA_COL[1])
        for records4 in records.values():
            for record in records4:
                # fill in the sheet
                row = ExcelHelper._row_offset_in_raw_data_sheet(mode=record.mode, seqs=seqs, qps=qps,
                                                                seq=record.name, qp=record.qp)
                for col, key in enumerate(PatKey.summary_patterns() + PatKey.summary_patterns_dec()):
                    sheet.cell(row, column=col + col_start, value=record[key])

    @staticmethod
    def _copy_cell_style(target_cell: Cell, source_cell: Cell):
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.style = copy(source_cell.style)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.fill = copy(source_cell.fill)
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
        if source_cell.hyperlink:
            target_cell.hyperlink(copy(source_cell.hyperlink))

        if source_cell.comment:
            target_cell.comment(copy(source_cell.comment))

    @staticmethod
    def _row_offset_in_raw_data_sheet(mode: Mode, seqs: Sequence[str], qps: Sequence[int],
                                      seq: Optional[str] = None, qp: Optional[int] = None):
        """
        当前mode的第一个数据，在原始数据sheet中的行偏移值(1-based)
        :param mode: mode
        :param seqs: sequences list
        :param qps: qp list
        :param seqs: sequence
        :param qp: qp
        :return: offset of first data in the raw data sheet
        """
        mode_index = 0
        for m in Mode:
            if m == mode:
                break
            mode_index += 1
        offset = mode_index * len(seqs) * len(qps) + 1
        if seq is not None and qp is not None:
            _id = ExcelHelper.seq_id(seqs, seq)
            offset += _id * len(qps) + list(qps).index(qp)
        return offset

    @staticmethod
    def fill_ref_data_header(mode, seqs, qps, excel: Excel, template_sheet: Worksheet, index: Optional[int] = None):
        # 从模板sheet拷贝
        sheet = excel.copy_sheet(source=template_sheet, new_title=mode.value, index=index)

        # 在原始数据sheet中的行偏移值
        row_offset = ExcelHelper._row_offset_in_raw_data_sheet(mode, seqs, qps)

        # 在当前sheet中的行偏移值
        row_start = row = 3
        col_start = column_index_from_string(_RAW_DATA_COL[1])
        cols = len(PatKey.summary_patterns() + PatKey.summary_patterns_dec())
        for seq_id, seq in enumerate(seqs):
            for qp in qps:
                if (row - row_start) % len(qps) == 0:
                    m = re.match(r"(\w+)_(\d+x\d+.*)", seq)
                    if m:
                        sheet.cell(row=row, column=column_index_from_string('A'), value=m.group(2))
                        sheet.cell(row=row, column=column_index_from_string('B'), value=m.group(1))
                    else:
                        sheet.cell(row=row, column=column_index_from_string('B'), value=seq)

                sheet.cell(row, column_index_from_string("C"), qp)
                for i, base_c in enumerate(['D', 'L']):
                    raw_sheet: Worksheet = excel[_RAW_DATA_SHEETS[i]]
                    base = column_index_from_string(base_c)
                    for j in range(cols):
                        raw_cell = raw_sheet.cell(row=row_offset, column=col_start + j)
                        sheet.cell(row, base + j, f"={raw_sheet.title}!{raw_cell.coordinate}")

                    # 将编码时间转化为小时为单位
                    cell: Cell = sheet.cell(row, base + cols)
                    cell.value = f"={cell.offset(row == 0, column=-2).coordinate}/3600"

                # 拷贝样式：向上偏移4行。注意：第一个序列所在的Cell的样式已从模板文件拷贝
                if seq_id != 0:
                    for col, _ in enumerate(sheet.columns):
                        cell = sheet.cell(row, col + 1)
                        ExcelHelper._copy_cell_style(cell, cell.offset(row=-4, column=0))

                row += 1
                row_offset += 1
