"""
此文件用于扫描编码日志，并提供编码的summary
具体工作方式为：

根据编码log的目录，遍历每个合法的编码log文件，编码log的合法性由filter_func来验证
然后打开该log文件，根本不同的编码器，读取summary信息，生成Record。

一个Record记录一个序列在特定QP下的Summary，也就是Excel表中的一行，通常一个序列有4个QP，也就是4个Record
整个文件夹中的全部log都会保存到records这个变量中，此结构为字典，键为序列名，值为该序列对应的4个Record

最终，将Record保存到指定的excel中或者输出到屏幕

"""
import os
import re
import openpyxl as op
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from abc import ABCMeta
from hpc.codec.mode import Mode
from typing import Optional


class Record:
    """
    关于编码信息的一个记录，用于填充BD-rate表
    """

    def __init__(self, _id: int, mode: Mode, name: str):
        """
        初始化一个记录
        :param _id: 当前记录的序列的ID
        :param mode: 当前记录所属的编码模式
        :param name: 当前记录所属的序列简称
        """
        self._id = _id
        self.mode = mode
        self.name = name
        self.qp = 0
        self.psnr_y = 0
        self.psnr_u = 0
        self.psnr_v = 0
        self.bitrate = 0
        self.encode_time = 0
        self.decode_time = 0
        # extra info
        self.frames = 0
        self.bits = 0
        self.ssim_y = 0

    def loc(self):
        """
        用于定位当前记录在excel中的行位置
        :return:
        """
        return "_".join([str(self.mode.value), str(self.name), str(self.qp)])

    def __str__(self):
        return ",".join([self.name, str(self.qp), str(self.bitrate), str(self.psnr_y), str(self.psnr_u), str(
            self.psnr_v), str(self.encode_time), str(self.decode_time)])

    def __repr__(self):
        return self.__str__()


class AbsLogScanner(metaclass=ABCMeta):
    def __init__(self,
                 enc_log_dir: str,
                 dec_log_dir: Optional[str],
                 seqs: list,
                 mode: Mode,
                 output_excel: str = None,
                 template: str = None,
                 is_anchor: bool = False,
                 is_separate: bool = False):
        """
        初始化一个日志扫描器
        :param enc_log_dir: 编码日志文件所在的目录
        :param dec_log_dir: 解码日志文件所在的目录
        :param seqs: 需要扫描的序列列表，序列的ID由序列在该列表的顺序确定
        :param mode: 当前编码的模式，用于确定在输出excel的位置
        :param output_excel: 输出的excel表格名称
        :param template: 输入的excel模板
        :param is_anchor: 是否是anchor数据，用于确定在excel中的填充位置
        :param is_separate: 是否为分片编码。TODO：HPM中待实现此功能
        """
        assert os.path.exists(enc_log_dir)
        self.enc_log_dir = enc_log_dir
        self.dec_log_dir = dec_log_dir
        self.seqs = seqs
        self.mode = mode
        self.out_excel = output_excel
        self.template = template
        self.is_anchor = is_anchor
        self.is_separate = is_separate
        self.records = dict()

    def _in_dict(self, file: str):
        for i, value in enumerate(self.seqs):
            if value in file:
                return i, value
        return None

    def _add_record(self, record: Record):
        """
        统一管理，将一个新的记录添加到成员变量中
        :param record: 新的记录
        """
        # 此处已经根据ID排序
        recodes_list = self.records.get(record._id) or list()
        recodes_list.append(record)
        recodes_list.sort(key=lambda r: r.qp)
        self.records[record._id] = recodes_list

    def _get_decode_time(self, dec_file: str, regex: str, enc_file: Optional[str] = None):
        """
        从解码文件中获取解码时间
        :param dec_file: 解码日志文件
        :param regex: 正则表达式，该表达式的group(1)指向解码时间字符串
        :param enc_file: 该解码文件对应的编码文件，如果不为None，则使用该文件验证一下二者所属的序列ID是否一致
        :return: 解码时间
        """
        dec_time = 0
        if self.dec_log_dir is not None and dec_file is not None and len(dec_file) > 0:
            if enc_file is not None:
                assert self._in_dict(dec_file) == self._in_dict(enc_file)
            with open(os.path.join(self.dec_log_dir, dec_file), "r") as fp:
                for line in fp:
                    m = re.match(regex, line)
                    if m:
                        dec_time = float(m.group(1))
                        break
        return dec_time

    def scan(self, filter_func_enc=None, filter_func_dec=None, rm_log=False):
        """
        执行扫描任务。从给定的文件夹中，挑选指定的文件进行解析
        :param filter_func_enc: 过滤掉不关心的编码日志文件
        :param filter_func_dec: 过滤掉不关心的解码日志文件
        :param rm_log: 扫描结束后是否删除log原始文件
        :return: 字典。key为所在序列的ID，value为当前序列对应的不同QP下的Record列表
        """
        raise NotImplemented

    def output(self):
        """
        将扫描到的记录全部输出到屏幕
        如果指定了template，同时会输出到Excel，并保存到out_excel, 如果未指定out_excel，则输出到以当前目录为名的excel表格中
        """
        if self.template is None or not os.path.exists(self.template):
            for name, records4 in self.records.items():
                records4 = sorted(records4, key=lambda r: int(r.qp))
                for record in records4:
                    print(record)
        else:
            sheet_names = ["Reference", "Test"]
            workbook: Workbook = op.load_workbook(self.template, keep_vba=True)
            sheet: Worksheet = workbook[sheet_names[0] if self.is_anchor else sheet_names[1]]
            for name, records4 in self.records.items():
                for record in records4:
                    record: Record = record
                    identifier = record.loc()
                    print(record)
                    # fill in the sheet
                    for index, value in enumerate(sheet.values):
                        if value[0] == identifier:
                            # both row and column indexes are started from 1 not 0
                            index += 1
                            s = 1
                            sheet.cell(index, column=s + 1, value=record.bitrate)
                            sheet.cell(index, column=s + 2, value=record.psnr_y)
                            sheet.cell(index, column=s + 3, value=record.psnr_u)
                            sheet.cell(index, column=s + 4, value=record.psnr_v)
                            sheet.cell(index, column=s + 5, value=record.encode_time)
                            break
            if self.out_excel is not None:
                workbook.save(self.out_excel)
            else:
                name = os.path.basename(os.path.abspath(os.curdir))
                workbook.save(name + "." + self.template.split(".")[-1])
            workbook.close()

    @staticmethod
    def get_valid_line_reg():
        """
        获取编码日志中，每一个编码帧日志输出的正则表达式，用于更新HPC的进度条
        :return:
        """
        raise NotImplemented

    @staticmethod
    def get_end_line_reg():
        """
        获取编码日志中，每一个序列编码完成时的正则表达式，用于结束HPC进度条的更新
        :return:
        """
        raise NotImplemented
