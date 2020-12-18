"""
此文件用于扫描编码日志，并提供编码的summary
具体工作方式为：

根据编码log的目录，遍历每个合法的编码log文件，编码log的合法性由filter_func来验证
然后打开该log文件，根本不同的编码器，读取summary信息，生成Record。

一个Record记录一个序列在特定QP下的Summary，也就是Excel表中的一行，通常一个序列有4个QP，也就是4个Record
整个文件夹中的全部log都会保存到records这个变量中，此结构为字典，键为序列名，值为该序列对应的4个Record

最终，将Record保存到指定的excel中或者输出到屏幕

"""
import os
import re
from hpc.core.helper import rmdir

import openpyxl as op
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Record:
    def __init__(self, _id: int, mode: str, name: str):
        self.id = _id
        self.mode = mode
        self.name = name
        self.qp = 0
        self.psnr_y = 0
        self.psnr_u = 0
        self.psnr_v = 0
        self.bitrate = 0
        self.time = 0

    def loc(self):
        return "_".join([str(self.mode), str(self.name), str(self.qp)])

    def __str__(self):
        return ",".join([self.name, str(self.qp), str(self.bitrate), str(self.psnr_y), str(self.psnr_u), str(
            self.psnr_v), str(self.time)])

    def __repr__(self):
        return self.__str__()


class _Scanner:
    def __init__(self, log_dir: str, seqs: list, mode: str,
                 output_excel: str = None,
                 template: str = None,
                 is_anchor: bool = False,
                 is_separate: bool = False):
        assert os.path.exists(log_dir)
        self.log_dir = log_dir
        self.seqs = seqs
        self.mode = mode
        self.out_excel = output_excel
        self.template = template
        self.is_anchor = is_anchor
        self.is_separate = is_separate
        self.records = dict()

    def _in_dict(self, file: str):
        for i, value in enumerate(self.seqs):
            if value in file:
                return i, value
        return None

    def _add_record(self, record: Record):
        # 此处已经根据ID排序
        if self.records.get(record.id) is None:
            self.records[record.id] = list()
        self.records[record.id].append(record)
        self.records[record.id] = sorted(self.records[record.id], key=lambda r: r.qp)

    def scan(self, filter_func=None, rm_log=False):
        raise NotImplemented

    def output(self):
        if self.template is None or not os.path.exists(self.template):
            for name, records4 in self.records.items():
                records4 = sorted(records4, key=lambda r: int(r.qp))
                for record in records4:
                    print(record)
        else:
            sheet_names = ["Reference", "Test"]
            workbook: Workbook = op.load_workbook(self.template, keep_vba=True)
            sheet: Worksheet = workbook[sheet_names[0] if self.is_anchor else sheet_names[1]]
            for name, records4 in self.records.items():
                for record in records4:
                    record: Record = record
                    print(record)
                    # fill in the sheet
                    for index, row in enumerate(sheet.rows):
                        index += 1
                        if row[0].value == record.loc():
                            s = 1
                            sheet.cell(index, column=s + 1, value=record.bitrate)
                            sheet.cell(index, column=s + 2, value=record.psnr_y)
                            sheet.cell(index, column=s + 3, value=record.psnr_u)
                            sheet.cell(index, column=s + 4, value=record.psnr_v)
                            sheet.cell(index, column=s + 5, value=record.time)
                            break
            if self.out_excel is not None:
                workbook.save(self.out_excel)
            else:
                name = os.path.basename(os.path.abspath(os.curdir))
                workbook.save(name + "." + self.template.split(".")[-1])
            workbook.close()


class HpmScanner(_Scanner):
    valid_line_reg = "\s*(\d+)\s+\(\s*[I^|P^|B]\)\s+(\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+)\s+(\d+)\s+(0\.\d+).+"
    end_line_reg = "Encoded\s*frame\s*count\s+=\s*(\d+)"

    def scan(self, filter_func=None, rm_log=False):
        files = os.listdir(self.log_dir)
        if callable(filter_func):
            files = filter(filter_func, files)
        # do merge first
        if self.is_separate:
            for seq in self.seqs:
                pass
        for file in files:
            _id, name = self._in_dict(file)
            assert name is not None
            if name is not None:
                record = Record(_id, self.mode, name)
                record.qp = int(file[-8:-6])
                with open(os.path.join(self.log_dir, file), "r") as fp:
                    for line in fp:
                        line = line.strip()
                        if "PSNR Y(dB)" in line:
                            record.psnr_y = float(line.strip("PSNR Y(dB)").strip().strip(":").strip())
                        elif "PSNR U(dB)" in line:
                            record.psnr_u = float(line.strip("PSNR U(dB)").strip().strip(":").strip())
                        elif "PSNR V(dB)" in line:
                            record.psnr_v = float(line.strip("PSNR V(dB)").strip().strip(":").strip())
                        elif "bitrate(kbps)" in line:
                            record.bitrate = float(line.strip("bitrate(kbps)").strip().strip(":").strip())
                        elif "Total encoding time" in line:
                            sp = line.strip("Total encoding time").strip().strip("=").strip().split(" ")
                            record.time = float(sp[2])
                            break
                self._add_record(record)
        if rm_log:
            rmdir(self.log_dir)
        self.records = dict(sorted(self.records.items(), key=lambda kv: kv[0]))
        return self.records


class Uavs3eScanner(HpmScanner):
    valid_line_reg = "\s*(\d+)\s*\(\s*[I|P|B]\)\|\s*(\d+\.\d+)\|\s*(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\|.+"
    end_line_reg = "Encoded\s*frame\s*count\s+=\s*(\d+)"

    def scan(self, filter_func=None, rm_log=False):
        file = os.path.join(self.log_dir, "psnr.txt")
        assert os.path.exists(file)
        with open(file, "r") as fp:
            for line in fp:
                matched = re.match(r"(\S+) {4}(\S+) (\S+) (\S+) (\S+) {4}(\S+) (\S+) (\S+) {4}(\S+)", line)
                assert matched is not None
                if matched:
                    file_name = matched.group(1)
                    _id, name = self._in_dict(file_name)
                    assert name is not None
                    record = Record(_id, self.mode, name)
                    record.qp = int(file_name[-6:-4])
                    record.bitrate = float(matched.group(2))
                    record.psnr_y = float(matched.group(3))
                    record.psnr_u = float(matched.group(4))
                    record.psnr_v = float(matched.group(5))
                    record.time = float(matched.group(9))
                    self._add_record(record)
        if rm_log:
            rmdir(file)
        self.records = dict(sorted(self.records.items(), key=lambda kv: kv[0]))
        return self.records


class Scanner(object):
    def __init__(self,
                 log_dir: str,
                 seqs: list,
                 mode: str,
                 scanner: str = None,
                 output_excel: str = None,
                 template: str = None,
                 is_anchor: bool = False,
                 is_parallel: bool = False):
        """
        :param log_dir: 日志目录
        :param seqs: 要扫描的日志名称简写
        :param mode: 模式，"AI"\"LDB"\"LDP"\"RA"之一，用于决定excel表中的sheet
        :param scanner: Scanner的名称，目前支持"HPM"\"UAVS3E"
        :param output_excel: 指定输出的excel表格的名称
        :param template: 指定excel表格模板的名称
        :param is_anchor: 指定当前数据是否是anchor，用于决定Excel中sheet
        :param is_parallel: 指定当前日志是否是并行编码的，仅适用于HPM分片编码
        """
        if scanner is None:
            scanner = HpmScanner.__name__.lower()
        if scanner.lower() in HpmScanner.__name__.lower():
            self.scanner: _Scanner = HpmScanner(log_dir, seqs, mode, output_excel, template, is_anchor, is_parallel)
        else:
            is_parallel = False
            self.scanner: _Scanner = Uavs3eScanner(log_dir, seqs, mode, output_excel, template, is_anchor, is_parallel)

    def scan(self, filter_func: callable = None, rm_log=False):
        """
        :param filter_func: 用于过滤文件夹中的非目标文件
        :param rm_log: 当扫描完成时，是否删除log文件
        :return: 字典列表
        """
        return self.scanner.scan(filter_func, rm_log)

    def output(self):
        self.scanner.output()
